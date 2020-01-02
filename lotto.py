import requests
from bs4 import BeautifulSoup 
from openpyxl import Workbook
import re
from openpyxl import load_workbook
import openpyxl
import os.path
import pandas as pd
import numpy as np
from openpyxl import workbook
from string import ascii_uppercase
from openpyxl.styles import Font, Color, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import DataBarRule, Color


data_bar_rule = DataBarRule(start_type='percent', start_value=0,end_type='percent',end_value=100,color=colors.RED)


thin_border = Border(left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

alpha_list = list(ascii_uppercase)

def main():
    file = 'lotto.xlsx'
    if os.path.isfile(file):
        new_ball()
        lastrow()
        if int(keyword[6]) == last_row -1:
            print('최신버전입니다.')
        else:
            wb = openpyxl.load_workbook('lotto.xlsx')
            sheet = wb.active
            row = sheet.max_row
            print(row)
            print('업데이트를 시작합니다.')
            for j in range(row, int(keyword[6])+1): #
                basic_url = "https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo=" 
                resp = requests.get(basic_url + str(j)) 
                soup = BeautifulSoup(resp.text, "lxml") 
                line = str(soup.find("meta", {"id" : "desc", "name" : "description"})['content']) 

                begin = line.find("당첨번호")
                begin = line.find(" ", begin) + 1 
                end = line.find(".", begin) 
                numbers = line[begin:end] 
                print("당첨번호" + str(j) +"회" , numbers)
                split_num = re.split('[, +]',numbers)
                write_ws = wb.active
                write_ws['A'+str(j+1)] = str(j) + "회"

                write_ws['B'+str(j+1)] = int(split_num[0])
                write_ws['C'+str(j+1)] = int(split_num[1])
                write_ws['D'+str(j+1)] = int(split_num[2])
                write_ws['E'+str(j+1)] = int(split_num[3])
                write_ws['F'+str(j+1)] = int(split_num[4])
                write_ws['G'+str(j+1)] = int(split_num[5])
                write_ws['H'+str(j+1)] = int(split_num[6])
            write_ws['H1'] = "추가번호"
            write_ws['B1'] = "1번째 번호"
            write_ws['C1'] = "2번째 번호"
            write_ws['D1'] = "3번째 번호"
            write_ws['E1'] = "4번째 번호"
            write_ws['F1'] = "5번째 번호"
            write_ws['G1'] = "6번째 번호"

            wb.save('lotto.xlsx')
            print('업데이트 완료')
    else:
        dowmload()

def dowmload():
    new_ball()
    print("다운로드 시작합니다.")
    basic_url = "https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo=" 
    write_wb = Workbook()
    mk_sheet = write_wb.active
    mk_sheet.title = '당첨번호 모음'
    for i in range(1,int(keyword[6])+1):#
        resp = requests.get(basic_url + str(i)) 
        soup = BeautifulSoup(resp.text, "lxml") 
        line = str(soup.find("meta", {"id" : "desc", "name" : "description"})['content']) 

        begin = line.find("당첨번호")
        begin = line.find(" ", begin) + 1 
        end = line.find(".", begin) 
        numbers = line[begin:end] 
        print("당첨번호" + str(i) +"회" , numbers)
        split_num = re.split('[, +]',numbers)

        if i == 0:
            pass
        else:
            mk_sheet['A'+str(i+1)] = str(i) + "회"

            mk_sheet['B'+str(i+1)] = int(split_num[0])
            mk_sheet['C'+str(i+1)] = int(split_num[1])
            mk_sheet['D'+str(i+1)] = int(split_num[2])
            mk_sheet['E'+str(i+1)] = int(split_num[3])
            mk_sheet['F'+str(i+1)] = int(split_num[4])
            mk_sheet['G'+str(i+1)] = int(split_num[5])
            mk_sheet['H'+str(i+1)] = int(split_num[6])
    
    mk_sheet['B1'] = "1번째 번호"
    mk_sheet['C1'] = "2번째 번호"
    mk_sheet['D1'] = "3번째 번호"
    mk_sheet['E1'] = "4번째 번호"
    mk_sheet['F1'] = "5번째 번호"
    mk_sheet['G1'] = "6번째 번호"
    mk_sheet['H1'] = "추가번호"   

    write_wb.save('lotto.xlsx')

    sheet2()
    print('다운로드 완료')

def sheet2():
    load_excel = openpyxl.load_workbook('lotto.xlsx')
    sheet2 = load_excel.create_sheet()
    sheet2.title = '데이터 분석'
    

    source = load_excel["당첨번호 모음"]
    count = 0
    # for cell in source['A']:
    #     print(cell).size
    for i in range(1,46):
        sheet2['A'+str(i+2)] = i

    sheet2['B1'] = "1번째 번호 갯수"
    sheet2['C1'] = "2번째 번호 갯수"
    sheet2['D1'] = "3번째 번호 갯수"
    sheet2['E1'] = "4번째 번호 갯수"
    sheet2['F1'] = "5번째 번호 갯수"
    sheet2['G1'] = "6번째 번호 갯수"
    sheet2['H1'] = "보너스 번호 갯수"
    for j in range(1,8):
        cell = sheet2.cell(row=2,column=j+1)
        cell.value = "=COUNT('당첨번호 모음'!{}:{})".format(alpha_list[j],alpha_list[j])
        cell.font = Font(bold=10)
        cell.border = thin_border

    for n in range(1,8):
        try:
            for m in range(1,46):
                cell = sheet2.cell(row=2+m, column=n+1)
                cell.value = "=COUNTIF('당첨번호 모음'!{}:{},{})/{}2".format(alpha_list[1+count],alpha_list[1+count],m,alpha_list[n])
                cell.font = Font(size=9)
                cell.border = thin_border
                cell.number_format = '0.####%'
            count += 1
        except IndexError:
            pass

    sheet2.conditional_formatting.add("B3:H47",data_bar_rule)
    load_excel.save('lotto.xlsx')
        
def lastrow():
    global last_row
    wb = openpyxl.load_workbook('lotto.xlsx')
    sheet = wb.active
    last_row = sheet.max_row


def new_ball():
    global keyword
    url = 'https://www.nlotto.co.kr/common.do?method=main'
    a = requests.get(url)
    soup = BeautifulSoup(a.text, "lxml")
    find = soup.find(id="lottoDrwNo")
    keyword = re.split('[< = " >]',str(find))

#1 / 8,145,060


if __name__ == "__main__":
    main()